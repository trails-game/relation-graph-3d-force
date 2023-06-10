from openpyxl.reader.excel import load_workbook
import threading
import requests
import json
import os

SEARCH_URL = "https://trails-game.com/wp-json/wp/v2/search"
BASE_URL_CHAR = "https://trails-game.com/character?p="
BASE_URL_ORG = "https://trails-game.com/org?p="

TYPES = ["Char", "Org", "Fam"]

def excel_to_dict(sheet, colNum):
   values = []
   # Calc the actual row number
   rowNum = len([row for row in sheet if not all([cell.value == None for cell in row])])
   
   for row in sheet.iter_rows(min_row=2, max_row=rowNum, max_col=colNum):
      dict = {}
      for cell in row:
         dict[sheet[cell.coordinate[0] + '1'].value] = cell.value
      values.append(dict)
         
   return values

def request_header(v, failed_links):
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.63 Safari/537.36 Edg/102.0.1245.33"}
    result = requests.head(str(v["avatar"]), headers=headers)
    if result.status_code >= 400:
        failed_links[v["name"]] = str(v["avatar"])

def search_for_link(name, new_node, type_):
    result = None
    if type_.split(".")[0] == "Char":
        result = requests.get(SEARCH_URL, 
        params={"type":"post", 
        "subtype": "dt_team", 
        "per_page":"1",
        "search": name})
    else:
        result = requests.get(SEARCH_URL, 
        params={"type":"post", 
        "subtype": "map", 
        "per_page":"1",
        "search": name})
    
    if result is not None and result.status_code == 200:
        responseJson = result.json()
        if len(responseJson) > 0:
            new_node["wikiPage"] = responseJson[0]["url"]
        else:
            new_node["wikiPage"] = ""
    else:
        new_node["wikiPage"] = ""

def parse_name_page(sheet, names, name_id_map, thread_list, malformed_types, failed_links, nodes):
    #name sheet processing
    id = 0
    values = excel_to_dict(sheet["角色"], 5)
    for v in values:
        if not v["name"] in names:
            names.add(v["name"])

            new_node = {"name" : v["name"], "id": str(id)}
            name_id_map[v["name"]] = str(id)

            id = id + 1
            if (v["avatar"] is not None) and v["avatar"].strip() != "":
                new_node["avatar"] = str(v["avatar"])
                t = threading.Thread(target=request_header, args=(v, failed_links))
                thread_list.append(t)
                t.start()

            if v["wikiPage"] is not None:
                new_node["wikiPage"] = v["wikiPage"]
            elif v["postid"] is not None and v["type"] == TYPES[0]:
                new_node["wikiPage"] = BASE_URL_CHAR + str(int(v["postid"]))
            elif v["postid"] is not None:
                new_node["wikiPage"] = BASE_URL_ORG + str(int(v["postid"]))
            else:
                t = threading.Thread(target=search_for_link, args=(v["name"], new_node, v["type"]))
                thread_list.append(t)
                t.start()

            new_node["type"]=v["type"] if v["type"] in TYPES else malformed_types.append(v)

            nodes.append(new_node)

def parse_relations(sheet, names, malformed_relations, missing_names, name_id_map, links):
    # not used
    # set的效率比list高很多。
    exising_src_dest_pairs = set()
    _values = excel_to_dict(sheet["人物组织关系"], 4)

    for v in _values:
        # 验证
        if (v["source"] == None or v["target"] == None or v["Relation"] == None or v["RelationType"] == None):
            malformed_relations.append(v)
            continue
        if (not v["source"] in names and not v["source"] in missing_names):
            missing_names.append(v["source"])
            continue
        if (not v["target"] in names and not v["target"] in missing_names):
            missing_names.append(v["target"])
            continue

        if (not v["source"] in missing_names and not v["target"] in missing_names):
            source_id = name_id_map[v["source"]]
            target_id = name_id_map[v["target"]]

        # 这里要用tuple，用dict的话，in就只会参考source_id而忽略target_id
        pair = (source_id, target_id)
        if not pair in exising_src_dest_pairs:
            exising_src_dest_pairs.add(pair)
            new_link = {"source":source_id, "target":target_id, "relation":v["Relation"], "type":v["RelationType"]}
            links.append(new_link)

def check_values(missing_names, malformed_types, malformed_relations, failed_links):
    if (len(missing_names) > 0):
        print("missing names: {0}".format(missing_names))
    if (len(malformed_types) > 0):
        print("malformed types: {0}".format(malformed_types) )
    if (len(malformed_relations) > 0):
        print("malformed relations: {0}".format(malformed_relations))
    if (len(failed_links) > 0): 
        print("failed avatar links: {0}".format(failed_links))
    if (len(missing_names) + len (malformed_types) + len(malformed_relations) + len(failed_links) > 0):
        raise ValueError("value error")

def write_outputs(output):
    target_file = os.path.join("dist", "data.json")
    with open(target_file, "w") as f:
        output = json.dumps(output, sort_keys=True, indent=4, ensure_ascii=False)
        # print(output)
        f.write(output)
        f.flush()

def run():
    file=os.path.join(os.getcwd(), "tools", "relation.xlsx")

    output = {"nodes":[], "links":[]}

    names = set()
    name_id_map = {}

    missing_names = []
    malformed_types = []
    malformed_relations = []
    failed_links = {}

    thread_list = []

    sheet = load_workbook(file, read_only=True)
    parse_name_page(sheet, names, name_id_map, thread_list, malformed_types, failed_links, output["nodes"])
    parse_relations(sheet, names, malformed_relations, missing_names, name_id_map, output["links"])

    # no use?
    for t in thread_list:
        t.join()

    check_values(missing_names, malformed_types, malformed_relations, failed_links)
    write_outputs(output)


if __name__ == "__main__":
    run()
